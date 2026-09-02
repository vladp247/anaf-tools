"""Excel + CSV export for bulk analysis results.
Supports indicator selection — caller passes a list of indicator keys to include.
"""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

H_DARK  = "1A1A2E"; H_MID = "16213E"; ROW_ALT = "F8F9FF"; WHITE = "FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _hf():    return Font(name="Calibri", bold=True, size=10, color="FFFFFF")
def _bdr():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)
def _cl():    return Alignment(horizontal="left",  vertical="center", wrap_text=True)
def _cr():    return Alignment(horizontal="right", vertical="center")
def _ca():    return Alignment(horizontal="left",  vertical="center")

def _aw(ws, mn=8, mx=42, sample=200):
    for col in ws.columns:
        sampled = list(col)[:sample + 1]
        w = min(max(max((len(str(c.value or "")) for c in sampled), default=mn) + 2, mn), mx)
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

def _hrow(ws, row, ncols, bg=H_DARK):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = _hf(); c.fill = _fill(bg); c.alignment = _ca(); c.border = _bdr()

def _cvt(v, year, eur_rates):
    if not eur_rates or v is None: return v
    rate = eur_rates.get(year, 1)
    return round(v / rate, 2) if rate else v

def _cur_label(eur_rates): return "EUR" if eur_rates else "RON"


# ── Indicator catalogue ───────────────────────────────────────────────
# Each entry: key → (label_template, getter_fn)
# label_template uses {y} for year and {cur} for currency label
# getter_fn(fd, y, eur_rates) → value

def _INDICATORS(cur):
    """Return ordered dict of all available per-year indicators."""
    return {
        # P&L
        "revenue":        (f"{{y}} Revenue ({{cur}})",        lambda fd,y,r: _cvt(fd.get("net_turnover") or 0, y, r)),
        "total_revenue":  (f"{{y}} Total Revenue ({{cur}})",  lambda fd,y,r: _cvt(fd.get("total_revenue") or 0, y, r)),
        "gross_result":   (f"{{y}} Gross Result ({{cur}})",   lambda fd,y,r: _cvt((fd.get("gross_profit") or 0)-(fd.get("gross_loss") or 0), y, r)),
        "net_result":     (f"{{y}} Net Result ({{cur}})",     lambda fd,y,r: _cvt(fd.get("net_result") or (fd.get("net_profit",0) or 0)-(fd.get("net_loss",0) or 0), y, r)),
        "total_expenses": (f"{{y}} Total Expenses ({{cur}})", lambda fd,y,r: _cvt(fd.get("total_expenses") or 0, y, r)),
        "margin_pct":     (f"{{y}} Margin %",                 lambda fd,y,r: fd.get("profit_margin_pct")),
        # Balance sheet
        "total_assets":   (f"{{y}} Total Assets ({{cur}})",   lambda fd,y,r: _cvt((fd.get("fixed_assets") or 0)+(fd.get("current_assets") or 0), y, r)),
        "fixed_assets":   (f"{{y}} Fixed Assets ({{cur}})",   lambda fd,y,r: _cvt(fd.get("fixed_assets") or 0, y, r)),
        "current_assets": (f"{{y}} Current Assets ({{cur}})", lambda fd,y,r: _cvt(fd.get("current_assets") or 0, y, r)),
        "inventories":    (f"{{y}} Inventories ({{cur}})",    lambda fd,y,r: _cvt(fd.get("inventories") or 0, y, r)),
        "receivables":    (f"{{y}} Receivables ({{cur}})",    lambda fd,y,r: _cvt(fd.get("receivables") or 0, y, r)),
        "cash":           (f"{{y}} Cash ({{cur}})",           lambda fd,y,r: _cvt(fd.get("cash_and_bank") or 0, y, r)),
        "equity":         (f"{{y}} Equity ({{cur}})",         lambda fd,y,r: _cvt(fd.get("total_equity") or 0, y, r)),
        "paid_in_capital":(f"{{y}} Paid-in Capital ({{cur}})",lambda fd,y,r: _cvt(fd.get("paid_in_capital") or 0, y, r)),
        "liabilities":    (f"{{y}} Liabilities ({{cur}})",    lambda fd,y,r: _cvt(fd.get("liabilities") or 0, y, r)),
        # Other
        "employees":      (f"{{y}} Employees",                lambda fd,y,r: fd.get("avg_employees") or 0),
    }

# Default selection — what gets exported if no selection is passed
DEFAULT_INDICATORS = [
    "revenue", "gross_result", "net_result", "margin_pct",
    "equity", "liabilities", "employees",
]


def _fin_cols_and_rows(results, years, eur_rates, selected_indicators):
    """Build (headers, row_getter) for the selected financial indicators."""
    cur  = _cur_label(eur_rates)
    yrs  = sorted(years)
    cat  = _INDICATORS(cur)
    sel  = [k for k in selected_indicators if k in cat]

    headers = []
    for y in yrs:
        for key in sel:
            tmpl, _ = cat[key]
            headers.append(tmpl.replace("{y}", str(y)).replace("{cur}", cur))

    def get_fin_row(fins):
        row = []
        for y in yrs:
            fd = fins.get(y)
            for key in sel:
                _, getter = cat[key]
                row.append(getter(fd, y, eur_rates) if fd else "")
        return row

    return headers, get_fin_row, len(sel)


# ── Base company columns (always included) ────────────────────────────

def _base_row(res):
    co   = res.get("company") or {}
    fins = res.get("financials") or {}
    addr = co.get("address_registered_office") or {}
    stare_list = res.get("onrc_stare") or []
    stare_str  = "; ".join(s.get("denumire", "") for s in stare_list)
    yrs = sorted(fins.keys())
    caen_desc = next((fins[y].get("caen_description", "") for y in yrs if fins.get(y)), "")
    caen_code = next((str(fins[y].get("caen_code", "")) for y in yrs if fins.get(y)), co.get("caen_code", ""))

    rev_yrs = sorted(y for y in yrs if fins.get(y) and (fins[y].get("net_turnover") or 0) > 0)
    cagr_rev = cagr_net = cagr_period = ""
    if len(rev_yrs) >= 2:
        sy, ey = rev_yrs[0], rev_yrs[-1]; cagr_period = f"{sy}→{ey}"
        try:
            cr = round(((fins[ey]["net_turnover"]/fins[sy]["net_turnover"])**(1/(ey-sy))-1)*100, 2)
            cagr_rev = f"{cr:+.2f}%"
        except: pass
    net_yrs = sorted(y for y in yrs if fins.get(y) and (fins[y].get("net_result") or 0) > 0)
    if len(net_yrs) >= 2:
        sy, ey = net_yrs[0], net_yrs[-1]
        try:
            nv = lambda y: (fins[y].get("net_result") or (fins[y].get("net_profit",0) or 0)-(fins[y].get("net_loss",0) or 0))
            cn = round(((nv(ey)/nv(sy))**(1/(ey-sy))-1)*100, 2)
            cagr_net = f"{cn:+.2f}%"
        except: pass

    return [
        res.get("cui",""), co.get("name",""), co.get("registration_number",""),
        co.get("registration_date",""), co.get("legal_form",""),
        addr.get("county",""), addr.get("locality",""),
        caen_code, caen_desc,
        "Yes" if co.get("is_vat_registered") else "No",
        "Yes" if co.get("is_inactive") else "No",
        co.get("fiscal_authority",""), co.get("phone",""),
        stare_str, cagr_rev, cagr_net, cagr_period,
    ]

BASE_HEADERS = [
    "CUI","Company Name","Reg No.","Reg Date","Legal Form",
    "County","City","CAEN Code","CAEN Description",
    "VAT Registered","Inactive","Fiscal Authority","Phone",
    "ONRC Status","Rev CAGR %","Net CAGR %","CAGR Period",
]

ONRC_HEADERS = [
    "ONRC Forma Juridica","ONRC County","ONRC City","ONRC Address",
    "ONRC Reg Date","Parent Country","Website",
]

def _onrc_row(res):
    od = res.get("onrc_data") or {}
    street = " ".join(filter(None, [od.get("adr_den_strada",""), od.get("adr_nr_strada","")]))
    return [
        od.get("forma_juridica",""), od.get("adr_judet",""),
        od.get("adr_localitate",""), street,
        od.get("data_inmatriculare",""), od.get("tara_firma_mama",""),
        od.get("web",""),
    ]


# ── Excel export ──────────────────────────────────────────────────────

def export_excel(results, errors, analytics, years, eur_rates=None,
                 selected_indicators=None):
    if selected_indicators is None:
        selected_indicators = DEFAULT_INDICATORS
    wb = Workbook()
    wb.remove(wb.active)
    _build_results(wb, results, years, eur_rates, selected_indicators)
    _build_summary(wb, analytics, eur_rates)
    _build_top10(wb, analytics, eur_rates)
    if errors:
        _build_errors(wb, errors)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_results(wb, results, years, eur_rates, selected_indicators):
    ws = wb.create_sheet("Results")
    yrs = sorted(years)
    has_onrc    = any(r.get("onrc_data") for r in results)
    large_sheet = len(results) > 2000

    fin_headers, get_fin_row, n_fin = _fin_cols_and_rows(
        results, yrs, eur_rates, selected_indicators
    )
    onrc_cols = ONRC_HEADERS if has_onrc else []
    all_cols  = BASE_HEADERS + onrc_cols + fin_headers

    ws.append(all_cols)
    _hrow(ws, 1, len(all_cols))
    if has_onrc:
        for ci in range(len(BASE_HEADERS)+1, len(BASE_HEADERS)+len(onrc_cols)+1):
            ws.cell(row=1, column=ci).fill = _fill("0070BF")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, res in enumerate(results, 2):
        fins = res.get("financials") or {}
        row  = _base_row(res)
        if has_onrc: row += _onrc_row(res)
        row += get_fin_row(fins)
        ws.append(row)
        if not large_sheet:
            fill = _fill(ROW_ALT) if i % 2 == 0 else _fill(WHITE)
            for col in range(1, len(all_cols)+1):
                c = ws.cell(row=i, column=col)
                c.fill = fill; c.border = _bdr(); c.alignment = _cl()

    _aw(ws, sample=200)


def _build_summary(wb, analytics, eur_rates):
    ws  = wb.create_sheet("Summary")
    cur = _cur_label(eur_rates)

    def _cvt_a(v, eur_rates):
        if not eur_rates or v is None: return v
        rate = eur_rates.get(2023, 4.9465)
        return round(v / rate, 2) if rate else v

    rows = [
        ("Companies analysed", analytics.get("total_companies")),
        ("Years", ", ".join(str(y) for y in analytics.get("years",[]))),
        ("Anchor year", analytics.get("anchor_year")),
        ("", None),
        (f"Avg Revenue ({cur})", _cvt_a(analytics.get("revenue_stats",{}).get("avg"), eur_rates)),
        (f"Median Revenue ({cur})", _cvt_a(analytics.get("revenue_stats",{}).get("median"), eur_rates)),
        (f"Total Revenue ({cur})", _cvt_a(analytics.get("revenue_stats",{}).get("total"), eur_rates)),
        ("Profitable", analytics.get("profitability",{}).get("profitable")),
        ("Loss-making", analytics.get("profitability",{}).get("loss_making")),
        ("Profitable %", analytics.get("profitability",{}).get("profitable_pct")),
        ("Avg Employees", analytics.get("employee_stats",{}).get("avg")),
    ]
    ws.append(["Metric", "Value"])
    _hrow(ws, 1, 2)
    for r, (lbl, val) in enumerate(rows, 2):
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name="Calibri", bold=True, size=10); lc.border = _bdr()
        vc = ws.cell(row=r, column=2, value=val)
        vc.alignment = _cr(); vc.border = _bdr()
    _aw(ws)


def _build_top10(wb, analytics, eur_rates):
    ws  = wb.create_sheet("Top 10")
    cur = _cur_label(eur_rates)
    top = analytics.get("top10_revenue") or []
    ay  = analytics.get("anchor_year", "")
    cols = ["Rank","CUI","Company","Revenue","Net Result","Employees","Entity Type"]
    ws.append(cols)
    _hrow(ws, 1, len(cols))
    for i, r in enumerate(top, 1):
        def cv(v):
            if not eur_rates or not v: return v
            return round(v / eur_rates.get(ay, 1), 2)
        row = [i, r.get("cui"), r.get("name"), cv(r.get("rev")),
               cv(r.get("net_result")), r.get("employees"), r.get("entity_type","")]
        ws.append(row)
        fill = _fill(ROW_ALT) if i % 2 == 0 else _fill(WHITE)
        for col in range(1, len(cols)+1):
            c = ws.cell(row=i+1, column=col)
            c.fill = fill; c.border = _bdr(); c.alignment = _cl()
    _aw(ws, 14, 40)


def _build_errors(wb, errors):
    ws = wb.create_sheet("Errors")
    ws.append(["CUI","Name","Error","Type"])
    _hrow(ws, 1, 4)
    for i, e in enumerate(errors, 2):
        row = [e.get("cui",""), e.get("name",""), e.get("error",""), e.get("error_type","")]
        ws.append(row)
        fill = _fill(ROW_ALT) if i % 2 == 0 else _fill(WHITE)
        for col in range(1, 5):
            c = ws.cell(row=i, column=col)
            c.fill = fill; c.border = _bdr(); c.alignment = _cl()
    _aw(ws, 8, 20)


# ── CSV export ────────────────────────────────────────────────────────

def export_csv(results, years, eur_rates=None, selected_indicators=None):
    import csv
    if selected_indicators is None:
        selected_indicators = DEFAULT_INDICATORS
    yrs = sorted(years)
    has_onrc = any(r.get("onrc_data") for r in results)

    fin_headers, get_fin_row, _ = _fin_cols_and_rows(
        results, yrs, eur_rates, selected_indicators
    )
    onrc_cols = ONRC_HEADERS if has_onrc else []
    all_cols  = BASE_HEADERS + onrc_cols + fin_headers

    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(all_cols)

    for res in results:
        fins = res.get("financials") or {}
        row  = _base_row(res)
        if has_onrc: row += _onrc_row(res)
        row += get_fin_row(fins)
        w.writerow(row)

    return buf.getvalue().encode("utf-8-sig")
