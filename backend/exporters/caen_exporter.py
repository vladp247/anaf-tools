"""CAEN Analytics Excel Exporter — Light and Full variants.

Light export (export_caen_excel):
  Sheets: Summary · Top 10 · Year Trend
  Rows:   Top 10 companies only
  Size:   ~50–100 KB — fast, easy to share

Full export (export_caen_full_excel):
  Sheets: Summary · ALL Companies · Top 10 · Year Trend
  Rows:   Every company in the sector for the anchor year
  Size:   Can be large (tens of thousands of rows for major sectors)
"""
from __future__ import annotations
import io, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.utils.logger import get_logger

log = get_logger(__name__)

H_DARK, H_MID = "003791", "0096FF"
ROW_ALT, WHITE = "F0F7FF", "FFFFFF"
OK_COLOR, ERR_COLOR = "059669", "DC2626"
SRC_COLORS = {"uu": "E8F4FF", "bl_bs": "E8F0FF", "ir": "F0E8FF", "live": "FFF8E8"}


def _hf(bold=True, color="FFFFFF", sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)

def _fill(c): return PatternFill("solid", fgColor=c)

def _bdr():
    t = Side(style="thin", color="C8EBFF")
    return Border(left=t, right=t, top=t, bottom=t)

def _al(): return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def _ac(): return Alignment(horizontal="center", vertical="center")
def _ar(): return Alignment(horizontal="right",  vertical="center")

def _aw(ws, mn=10, mx=48):
    for col in ws.columns:
        mx_len = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx_len + 2, mn), mx)

def _hrow(ws, row_num, ncols, bg=H_DARK):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=col)
        c.font = _hf(); c.fill = _fill(bg); c.alignment = _ac(); c.border = _bdr()

def _cvt(v, year, eur_rates):
    if v is None or not eur_rates: return v
    rate = eur_rates.get(year, eur_rates.get(max(eur_rates), 5.0))
    try:    return round(v / rate, 2)
    except: return v

def _cur(eur_rates): return "EUR" if eur_rates else "RON"


# ══════════════════════════════════════════════════════════════════════
# LIGHT EXPORT  (Summary + Top10 + Trend)
# ══════════════════════════════════════════════════════════════════════
def export_caen_excel(result: dict, years: list[int], eur_rates: dict | None = None) -> bytes:
    """Light export: Summary + Top 10 + Year Trend. Fast, shareable."""
    cur = _cur(eur_rates)
    log.info("CAEN Light export: CAEN %s, %d years, %s", result.get("caen"), len(years), cur)
    wb = Workbook(); wb.remove(wb.active)
    _sheet_summary(wb, result, years, eur_rates, cur)
    _sheet_top10(wb, result, years, eur_rates, cur)
    _sheet_trend(wb, result, years, eur_rates, cur)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════
# FULL EXPORT  (Summary + ALL Companies + Top 10 + Trend)
# ══════════════════════════════════════════════════════════════════════
def export_caen_full_excel(
    result: dict,
    all_rows: list[dict],
    years: list[int],
    eur_rates: dict | None = None,
) -> bytes:
    """Full export: every company in sector. Can be large."""
    cur = _cur(eur_rates)
    log.info("CAEN Full export: CAEN %s, %d companies, %d years, %s",
             result.get("caen"), len(all_rows), len(years), cur)
    wb = Workbook(); wb.remove(wb.active)
    _sheet_summary(wb, result, years, eur_rates, cur, full=True)
    _sheet_all_companies(wb, result, all_rows, years, eur_rates, cur)
    _sheet_top10(wb, result, years, eur_rates, cur)
    _sheet_trend(wb, result, years, eur_rates, cur)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════
# Sheet helpers
# ══════════════════════════════════════════════════════════════════════

def _sheet_summary(wb, d, years, eur_rates, cur, full=False):
    ws = wb.create_sheet("Summary")
    anchor = d.get("anchor_year", max(years))

    def hdr(txt, r, c, span=1, bg=H_DARK):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = _hf(); cell.fill = _fill(bg); cell.alignment = _ac(); cell.border = _bdr()
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
            for cc in range(c+1, c+span):
                ws.cell(row=r, column=cc).fill = _fill(bg)

    def kv(lbl, val, r):
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name="Calibri", bold=True, size=10); lc.border = _bdr(); lc.alignment = _al()
        vc = ws.cell(row=r, column=2, value=val)
        vc.border = _bdr(); vc.alignment = _ar()

    def cvts(v): return _cvt(v, anchor, eur_rates)

    r = 1
    export_type = "Full Export — All Companies" if full else "Light Export — Top 10 Summary"
    hdr(f"CAEN {d.get('caen')} — {d.get('description','')}", r, 1, 2, H_DARK); r+=1
    hdr(f"Exported {datetime.date.today()} · {cur} · Anchor {anchor} · {export_type}", r, 1, 2, H_MID); r+=2

    hdr("Sector Overview", r, 1, 2); r+=1
    kv("CAEN Code",        d.get("caen"),             r); r+=1
    kv("Description",      d.get("description","—"),  r); r+=1
    kv("Anchor Year",      anchor,                    r); r+=1
    kv("Years Analyzed",   ", ".join(str(y) for y in sorted(years)), r); r+=1
    kv("Total Companies",  d.get("total_companies",0),r); r+=1

    src = d.get("source_breakdown") or {}
    kv("  Source: UU (micro/small)", src.get("uu",0),   r); r+=1
    kv("  Source: BL/BS (med/large)",src.get("bl_bs",0),r); r+=1
    kv("  Source: IFRS",             src.get("ir",0),   r); r+=1
    kv("  Source: Live ANAF scan",   src.get("live",0), r); r+=2

    rv  = d.get("revenue_stats")    or {}
    nt  = d.get("net_result_stats") or {}
    emp = d.get("employee_stats")   or {}

    hdr(f"Revenue ({cur}, {anchor})", r, 1, 2); r+=1
    for lbl, val in [("Avg",cvts(rv.get("avg"))),("Median",cvts(rv.get("median"))),
                     ("Total",cvts(rv.get("total"))),("Max",cvts(rv.get("max"))),("Min",cvts(rv.get("min")))]:
        kv(lbl, val, r); r+=1
    r+=1

    hdr(f"Net Result ({cur}, {anchor})", r, 1, 2); r+=1
    for lbl, val in [("Avg",cvts(nt.get("avg"))),("Median",cvts(nt.get("median"))),("Total",cvts(nt.get("total")))]:
        kv(lbl, val, r); r+=1
    r+=1

    hdr("Employee Stats", r, 1, 2); r+=1
    kv("Avg",    emp.get("avg"),   r); r+=1
    kv("Median", emp.get("median"),r); r+=1
    kv("Total",  emp.get("total"), r); r+=2

    prof = d.get("profitability") or {}
    hdr("Profitability", r, 1, 2); r+=1
    kv("Profitable",     prof.get("profitable",0),  r); r+=1
    kv("Loss-making",    prof.get("loss_making",0), r); r+=1
    kv("Breakeven",      prof.get("breakeven",0),   r); r+=1
    kv("Profitable %",   f"{prof.get('profitable_pct',0):.1f}%", r); r+=2

    pc = d.get("portfolio_cagr") or {}
    hdr(f"Sector CAGR ({pc.get('start_year','?')}→{pc.get('end_year','?')})", r, 1, 2); r+=1
    kv("Revenue CAGR (median %)", f"{pc.get('revenue_median','')}%" if pc.get("revenue_median") is not None else "—", r); r+=1
    kv("Revenue CAGR (avg %)",    f"{pc.get('revenue_avg','')}%"    if pc.get("revenue_avg")    is not None else "—", r); r+=1
    kv("Net CAGR (median %)",     f"{pc.get('net_median','')}%"     if pc.get("net_median")     is not None else "—", r); r+=1
    kv("Companies in CAGR calc",  pc.get("companies",0), r); r+=2

    sz = d.get("size_buckets") or {}
    hdr("Size Distribution (by Revenue)", r, 1, 2); r+=1
    for nm in ["Micro","Small","Medium","Large"]:
        kv(nm, sz.get(nm,0), r); r+=1
    r+=1

    pct = d.get("percentiles") or {}
    hdr(f"Revenue Percentiles ({cur})", r, 1, 2); r+=1
    for p in ["10","25","50","75","90","95","99"]:
        if p in pct: kv(f"p{p}", cvts(pct[p]), r); r+=1

    _aw(ws, 18, 40)


def _sheet_all_companies(wb, d, all_rows, years, eur_rates, cur):
    """Full company list — every company with financials for all selected years."""
    ws    = wb.create_sheet(f"All Companies ({len(all_rows):,})")
    yrs   = sorted(years)
    anchor = d.get("anchor_year", max(yrs))

    id_cols = ["Rank", "CUI", "Company Name", "Source", "Entity Type"]
    yr_cols = []
    for y in yrs:
        yr_cols += [
            f"{y} Revenue ({cur})",
            f"{y} Gross Result ({cur})",
            f"{y} Net Result ({cur})",
            f"{y} Employees",
        ]
    cols = id_cols + yr_cols
    ws.append(cols)
    _hrow(ws, 1, len(cols), H_DARK)
    for gi, y in enumerate(yrs):
        bg = "1A3A5C" if gi % 2 == 0 else H_MID
        start = len(id_cols) + gi * 4 + 1
        for ci in range(start, start + 4):
            ws.cell(row=1, column=ci).fill = _fill(bg)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for i, r in enumerate(all_rows, 1):
        src  = r.get("source", "uu")
        row_vals = [i, r.get("cui",""), r.get("name",""), src.upper(), r.get("entity_type","")]
        yr_data = r.get("years", {})
        for y in yrs:
            fd = yr_data.get(y) or {}
            row_vals += [
                _cvt(fd.get("rev") or 0,          y, eur_rates) if fd.get("rev")          is not None else "",
                _cvt(fd.get("gross_result") or 0,  y, eur_rates) if fd.get("gross_result") is not None else "",
                _cvt(fd.get("net_result") or 0,    y, eur_rates) if fd.get("net_result")   is not None else "",
                fd.get("employees") or ""          if fd.get("employees")   is not None else "",
            ]
        ws.append(row_vals)

        fill = _fill(SRC_COLORS.get(src, WHITE)) if i % 2 == 1 else _fill("F8FCFF")
        for col_idx in range(1, len(cols) + 1):
            c = ws.cell(row=i + 1, column=col_idx)
            c.fill = fill; c.border = _bdr(); c.alignment = _al()
            if col_idx > len(id_cols):
                offset = (col_idx - len(id_cols) - 1) % 4
                if offset in (0, 1, 2):
                    c.alignment = _ar(); c.number_format = '#,##0'
                if offset == 2:
                    try:
                        val_yr_idx = (col_idx - len(id_cols) - 1) // 4
                        val_y = yrs[val_yr_idx]
                        net = (r.get("years", {}).get(val_y) or {}).get("net_result", 0) or 0
                        c.font = Font(name="Calibri", size=10, bold=True,
                                      color=OK_COLOR if net >= 0 else ERR_COLOR)
                    except Exception:
                        pass
        if i % 5000 == 0:
            log.info("CAEN full export: wrote %d/%d rows", i, len(all_rows))

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    for gi in range(len(yrs)):
        for offset, width in [(0, 20), (1, 18), (2, 18), (3, 11)]:
            cl = get_column_letter(len(id_cols) + gi * 4 + offset + 1)
            ws.column_dimensions[cl].width = width

    note_row = len(all_rows) + 3
    note = ws.cell(row=note_row, column=1,
                   value=(f"CAEN {d.get('caen')} — {len(all_rows):,} companies — "
                          f"Years: {', '.join(str(y) for y in yrs)} — "
                          f"Exported {datetime.date.today()}"))
    note.font = Font(name="Calibri", italic=True, color="666666", size=9)
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=min(len(cols), 12))


def _sheet_top10(wb, d, years, eur_rates, cur):
    ws = wb.create_sheet("Top 10 by Revenue")
    anchor = d.get("anchor_year", max(years))
    top10  = d.get("top10_revenue") or []

    ws.cell(row=1, column=1, value=f"Top 10 — CAEN {d.get('caen')} — {cur} — Anchor {anchor}")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=12, color=H_DARK)
    ws.merge_cells("A1:I1")

    cols = ["Rank","CUI","Company Name",f"Revenue ({cur})",f"Net Result ({cur})",
            "Employees","Margin %",f"Total Assets ({cur})",f"Equity ({cur})"]
    ws.append(cols); _hrow(ws,2,len(cols),H_MID)
    ws.row_dimensions[2].height=22; ws.freeze_panes="A3"

    for i, r in enumerate(top10, 1):
        rev  = _cvt(r.get("rev",0),         anchor, eur_rates)
        net  = _cvt(r.get("net_result",0),  anchor, eur_rates)
        ta   = _cvt(r.get("total_assets",0),anchor, eur_rates)
        eq   = _cvt(r.get("equity",0),      anchor, eur_rates)
        margin = round(net/rev*100,1) if rev else None
        ws.append([i, r.get("cui",""), r.get("name",""), rev, net,
                   r.get("employees",0), margin, ta, eq])
        fill = _fill(ROW_ALT) if i%2==0 else _fill(WHITE)
        for col in range(1, len(cols)+1):
            c = ws.cell(row=i+2, column=col)
            c.fill=fill; c.border=_bdr(); c.alignment=_al()
            if col in (4,5,8,9): c.alignment=_ar(); c.number_format='#,##0'
            if col==5:
                net_val = r.get("net_result",0) or 0
                c.font=Font(name="Calibri",size=10,bold=True,
                            color=OK_COLOR if net_val>=0 else ERR_COLOR)

    ws.column_dimensions["C"].width = 40
    for cl in ["D","E","F","G","H","I"]: ws.column_dimensions[cl].width = 18


def _sheet_trend(wb, d, years, eur_rates, cur):
    ws = wb.create_sheet("Year Trend")
    trend = d.get("trend") or []

    ws.cell(row=1,column=1,value=f"CAEN {d.get('caen')} — Year-by-Year Trend — {cur}")
    ws.cell(row=1,column=1).font=Font(name="Calibri",bold=True,size=12,color=H_DARK)
    ws.merge_cells("A1:H1")

    cols=["Year","Companies",f"Avg Revenue ({cur})",f"Median Revenue ({cur})",
          f"Total Revenue ({cur})",f"Avg Net Result ({cur})","Profitable Count","Profitable %"]
    ws.append(cols); _hrow(ws,2,len(cols),H_MID)
    ws.row_dimensions[2].height=22

    for i,t in enumerate(trend,1):
        y=t.get("year",0)
        ws.append([
            y, t.get("companies",0),
            _cvt(t.get("avg_rev"),    y, eur_rates),
            _cvt(t.get("median_rev"), y, eur_rates),
            _cvt(t.get("total_rev"),  y, eur_rates),
            _cvt(t.get("avg_net"),    y, eur_rates),
            t.get("profitable",0),
            round(t["profitable"]/t["companies"]*100,1)
                if t.get("companies") and t.get("profitable") is not None else None,
        ])
        fill=_fill(ROW_ALT) if i%2==0 else _fill(WHITE)
        for col in range(1,len(cols)+1):
            c=ws.cell(row=i+2,column=col)
            c.fill=fill; c.border=_bdr(); c.alignment=_al()

    _aw(ws,12,28)
