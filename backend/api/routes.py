"""FastAPI routes — all HTTP endpoints."""
from __future__ import annotations
import asyncio
import io
import csv
import datetime
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel

from config import Config
from backend.api.anaf_client import get_anaf_client
from backend.services.lookup_service import lookup_single
from backend.services.bulk_processor import run_bulk_job, retry_failed
from backend.services.analytics import compute_analytics
from backend.exporters.excel_exporter import export_excel, export_csv
from backend.validators.validators import validate_cui, validate_years, validate_and_parse_csv
from backend.jobs.job_manager import get_job_manager
from backend.services.onrc_service import get_onrc, get_index_state
from backend.services.caen_service import get_caen, get_index_state as caen_index_state
from backend.services.name_matcher import (
    get_match_manager, run_match_job, normalize_name,
    AUTO_THRESHOLD, REVIEW_THRESHOLD,
)
from backend.utils.logger import get_logger

log = get_logger(__name__)

FRONTEND = Path(__file__).parent.parent.parent / "frontend" / "index.html"

app = FastAPI(title="ANAF Intel", version="2.1.0")


@app.on_event("startup")
async def startup():
    log.info("ANAF Intel started — http://%s:%d", Config.HOST, Config.PORT)
    try:
        get_onrc().load_nomenclator()
    except Exception:
        pass


@app.on_event("shutdown")
async def shutdown():
    await get_anaf_client().close()


# ── Frontend ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root():
    if not FRONTEND.exists():
        raise HTTPException(500, "Frontend not found")
    return HTMLResponse(FRONTEND.read_text(encoding="utf-8"))


# ── Setup / first-run ─────────────────────────────────────────────────────────

@app.get("/api/setup/status")
async def setup_status():
    first_run = not Config.SETUP_DONE_PATH.exists()

    # CSV age
    csv_date_str = None
    csv_age_days = None
    csv_stale    = False
    if Config.CSV_VERSION_PATH.exists():
        raw = Config.CSV_VERSION_PATH.read_text().strip()
        csv_date_str = raw
        try:
            # Accept YYYY-MM or YYYY-MM-DD
            parts = raw.split("-")
            if len(parts) == 2:
                ref = datetime.date(int(parts[0]), int(parts[1]), 1)
            else:
                ref = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
            csv_age_days = (datetime.date.today() - ref).days
            csv_stale    = csv_age_days > 60
        except Exception:
            pass

    svc    = get_onrc()
    files  = svc.files_status()
    indexed = svc.is_indexed()
    stats  = svc.db_stats() if indexed else {}

    return JSONResponse({
        "first_run":     first_run,
        "csv_date":      csv_date_str,
        "csv_age_days":  csv_age_days,
        "csv_stale":     csv_stale,
        "files":         files,
        "indexed":       indexed,
        "stats":         stats,
        "index_state":   get_index_state(),
    })


@app.post("/api/setup/complete")
async def setup_complete():
    Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    Config.SETUP_DONE_PATH.write_text("done")
    return {"ok": True}


# ── EUR rates ─────────────────────────────────────────────────────────────────

@app.get("/api/eur-rates")
async def eur_rates():
    return JSONResponse(Config.EUR_RATES)


# ── ONRC ──────────────────────────────────────────────────────────────────────

@app.get("/api/onrc/files-debug")
async def onrc_files_debug():
    """Shows exact resolved paths for each expected file — useful for diagnosing detection issues."""
    from backend.services.onrc_service import _resolve_path
    paths = {
        "firme":       (Config.ONRC_FIRME_PATH, _resolve_path(Config.ONRC_FIRME_PATH)),
        "stare":       (Config.ONRC_STARE_PATH, _resolve_path(Config.ONRC_STARE_PATH)),
        "reps":        (Config.ONRC_REPS_PATH,  _resolve_path(Config.ONRC_REPS_PATH)),
        "nomenclator": (Config.NOMENCLATOR_PATH, _resolve_path(Config.NOMENCLATOR_PATH)),
    }
    return JSONResponse({
        k: {
            "expected": str(expected),
            "resolved": str(resolved),
            "exists":   resolved.exists(),
            "match":    expected.name != resolved.name,
        }
        for k, (expected, resolved) in paths.items()
    })


@app.get("/api/onrc/status")
async def onrc_status():
    svc    = get_onrc()
    files  = svc.files_status()
    indexed = svc.is_indexed()
    stats  = svc.db_stats() if indexed else {}
    return JSONResponse({
        "files":       files,
        "indexed":     indexed,
        "index_state": get_index_state(),
        "stats":       stats,
    })


@app.post("/api/onrc/start-index")
async def onrc_start_index():
    svc   = get_onrc()
    files = svc.files_status()
    if not files["all_csv"]:
        raise HTTPException(400, "Required CSV files not found in data/ folder")
    started = svc.start_indexing()
    return {"started": started, "message": "Indexing started" if started else "Already indexing"}


@app.get("/api/onrc/index-progress")
async def onrc_progress():
    return JSONResponse(get_index_state())


@app.get("/api/search/name")
async def search_name(q: str = "", limit: int = 20):
    if not q or len(q) < 2:
        return JSONResponse([])
    results = get_onrc().search_by_name(q.strip(), min(limit, 50))
    return JSONResponse(results)


# ── Template ──────────────────────────────────────────────────────────────────

@app.get("/api/template")
async def csv_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["CUI"])
    for cui in ["14399840", "6480536", "40066640"]:
        w.writerow([cui])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=anaf_cuis.csv"},
    )


# ── Check ─────────────────────────────────────────────────────────────────────

class CheckReq(BaseModel):
    cui: str
    years: list[int]


@app.post("/api/check")
async def check(req: CheckReq):
    ok, msg, cui = validate_cui(req.cui)
    if not ok:
        raise HTTPException(422, msg)
    ok2, msg2 = validate_years(req.years)
    if not ok2:
        raise HTTPException(422, msg2)

    result = await lookup_single(cui, req.years)

    try:
        svc = get_onrc()
        if svc.is_indexed():
            result = svc.enrich(result)
    except Exception as ex:
        log.warning("ONRC enrich failed: %s", ex)

    return JSONResponse(content=result)


# ── Bulk ──────────────────────────────────────────────────────────────────────

@app.post("/api/bulk/start")
async def bulk_start(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    years: str = Form("2022,2023"),
    onrc_enrich: str = Form("false"),
    offline_financials: str = Form("false"),
):
    try:
        year_list = [int(y.strip()) for y in years.split(",") if y.strip()]
    except ValueError:
        raise HTTPException(422, "Invalid years format")
    ok_y, msg_y = validate_years(year_list)
    if not ok_y:
        raise HTTPException(422, msg_y)

    content = await file.read()
    ok_csv, msg_csv, raw_cuis, warns = validate_and_parse_csv(content)
    if not ok_csv:
        raise HTTPException(422, msg_csv)

    seen: set[int] = set()
    unique: list[int] = []
    for c in raw_cuis:
        if c not in seen:
            seen.add(c)
            unique.append(c)
    dupes = len(raw_cuis) - len(unique)

    jm  = get_job_manager()
    job = jm.create(unique, year_list)
    job.original_count     = len(raw_cuis)
    job.duplicates_removed = dupes
    job.invalid_removed    = len(warns)
    job.onrc_enrich        = onrc_enrich.lower() in ("true", "1", "yes")
    job.offline_financials = offline_financials.lower() in ("true", "1", "yes")
    for w in warns[:10]:
        job.add_log(f"  ⚠ {w}")
    job.add_log(f"CSV: {job.original_count} rows → {len(unique)} unique ({dupes} dupes removed)")

    task = asyncio.create_task(run_bulk_job(job))
    job._task = task

    return JSONResponse({
        "job_id":            job.job_id,
        "total":             job.total,
        "duplicates_removed": dupes,
        "invalid_removed":   len(warns),
        "years":             year_list,
        "warnings":          warns[:20],
    })


@app.get("/api/bulk/status/{jid}")
async def bulk_status(jid: str):
    job = get_job_manager().get(jid)
    if not job:
        raise HTTPException(404, f"Job '{jid}' not found")
    return JSONResponse(job.to_dict())


@app.post("/api/bulk/pause/{jid}")
async def bulk_pause(jid: str):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    if job.status != "running": raise HTTPException(400, f"Job is {job.status}")
    job.pause()
    return {"ok": True, "status": job.status}


@app.post("/api/bulk/resume/{jid}")
async def bulk_resume(jid: str):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    if job.status != "paused": raise HTTPException(400, f"Job is {job.status}")
    job.resume()
    return {"ok": True, "status": job.status}


@app.post("/api/bulk/cancel/{jid}")
async def bulk_cancel(jid: str):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    job.cancel()
    return {"ok": True, "status": job.status}


@app.post("/api/bulk/retry-failed/{jid}")
async def bulk_retry(jid: str):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    if job.status not in ("done", "cancelled", "error"):
        raise HTTPException(400, "Job must be done/cancelled before retrying")
    if not job.errors:
        return {"ok": True, "retrying": 0, "message": "No errors to retry"}
    task = asyncio.create_task(retry_failed(job))
    job._task = task
    return {"ok": True, "retrying": len(job.errors)}


@app.get("/api/bulk/year-coverage/{jid}")
async def bulk_year_coverage(jid: str, year: int = Query(...)):
    """Returns how many companies have/don't have financials for a given year."""
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    have, missing, examples = [], [], []
    for r in job.results:
        fins = r.get("financials") or {}
        fd   = fins.get(year)
        rev  = (fd or {}).get("net_turnover") or 0
        co   = r.get("company") or {}
        name = co.get("name") or f"CUI {r.get('cui','')}"
        if fd and rev > 0:
            have.append(r.get("cui"))
        else:
            missing.append(r.get("cui"))
            if len(examples) < 5:
                examples.append(name)
    return JSONResponse({
        "year":         year,
        "total":        len(job.results),
        "have_data":    len(have),
        "missing":      len(missing),
        "examples":     examples,
        "missing_cuis": missing,
    })


@app.get("/api/bulk/analytics/{jid}")
async def bulk_analytics(
    jid: str,
    anchor_year:  int = Query(default=0),
    require_year: int = Query(default=0),   # 0 = no filter; >0 = only companies with this year
):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    results = job.results
    if require_year:
        results = [r for r in results
                   if (r.get("financials") or {}).get(require_year)
                   and ((r.get("financials") or {}).get(require_year, {}) or {}).get("net_turnover", 0)]
    ay = anchor_year if anchor_year in job.years else None
    return JSONResponse(compute_analytics(results, job.years, anchor_year=ay))


class ExportConfig(BaseModel):
    indicators:      list[str] | None = None
    years:           list[int] | None = None
    currency:        str = "RON"
    fmt:             str = "xlsx"
    include_onrc:    bool = True
    require_year:    int | None = None   # only include companies with data for this year
    anchor_override: int | None = None   # override anchor year (e.g. 2024 when 2025 unavailable)


@app.post("/api/bulk/export-custom/{jid}")
async def bulk_export_custom(jid: str, cfg: ExportConfig):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")

    eur_rates  = Config.EUR_RATES if cfg.currency == "EUR" else None
    years      = sorted(cfg.years) if cfg.years else job.years
    indicators = cfg.indicators
    anchor     = cfg.anchor_override or None

    # Filter results
    results = job.results
    if cfg.require_year:
        results = [r for r in results
                   if (r.get("financials") or {}).get(cfg.require_year)
                   and ((r.get("financials") or {}).get(cfg.require_year, {}) or {}).get("net_turnover", 0)]
    if years != sorted(job.years):
        results = [{**r, "financials": {y: (r.get("financials") or {}).get(y) for y in years}} for r in results]
    if not cfg.include_onrc:
        results = [{**r, "onrc_data": None} for r in results]

    loop = asyncio.get_event_loop()
    date = datetime.date.today().strftime("%Y%m%d")
    cur  = cfg.currency

    if cfg.fmt == "csv":
        csv_bytes = await loop.run_in_executor(
            None, lambda: export_csv(results, years, eur_rates=eur_rates,
                                     selected_indicators=indicators)
        )
        return StreamingResponse(
            io.BytesIO(csv_bytes), media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename=anaf_bulk_{jid}_{cur}_{date}.csv"},
        )
    else:
        def _build():
            analytics = compute_analytics(results, years, anchor_year=anchor)
            return export_excel(results, job.errors, analytics, years,
                                eur_rates=eur_rates, selected_indicators=indicators)
        xlsx = await loop.run_in_executor(None, _build)
        return StreamingResponse(
            io.BytesIO(xlsx),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=anaf_bulk_{jid}_{cur}_{date}.xlsx"},
        )


@app.get("/api/bulk/export/{jid}")
async def bulk_export(jid: str, eur: int = Query(default=0),
                      indicators: str = Query(default="")):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    eur_rates = Config.EUR_RATES if eur else None
    sel  = [i.strip() for i in indicators.split(",") if i.strip()] or None
    loop = asyncio.get_event_loop()
    def _build():
        analytics = compute_analytics(job.results, job.years)
        return export_excel(job.results, job.errors, analytics, job.years,
                            eur_rates=eur_rates, selected_indicators=sel)
    xlsx = await loop.run_in_executor(None, _build)
    date     = datetime.date.today().strftime("%Y%m%d")
    currency = "EUR" if eur else "RON"
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=anaf_bulk_{jid}_{currency}_{date}.xlsx"},
    )


@app.get("/api/bulk/export-csv/{jid}")
async def bulk_export_csv(jid: str, eur: int = Query(default=0),
                          indicators: str = Query(default="")):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    eur_rates = Config.EUR_RATES if eur else None
    sel  = [i.strip() for i in indicators.split(",") if i.strip()] or None
    loop = asyncio.get_event_loop()
    csv_bytes = await loop.run_in_executor(
        None, lambda: export_csv(job.results, job.years,
                                 eur_rates=eur_rates, selected_indicators=sel)
    )
    date     = datetime.date.today().strftime("%Y%m%d")
    currency = "EUR" if eur else "RON"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename=anaf_bulk_{jid}_{currency}_{date}.csv"},
    )


@app.get("/api/bulk/error-log/{jid}")
async def bulk_error_log(jid: str):
    job = get_job_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["CUI", "Name", "Error Type", "Error"])
    for e in job.errors:
        w.writerow([e.get("cui", ""), e.get("name", ""), e.get("error_type", ""), e.get("error", "")])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=anaf_errors_{jid}.csv"},
    )


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "2.1"}


# ── CAEN Analytics ────────────────────────────────────────────────────

@app.get("/api/caen/status")
async def caen_status():
    svc  = get_caen()
    return JSONResponse({
        "files_detail":  svc.files_status_detail(),
        "years_available": svc.years_available(),
        "indexed":       svc.is_indexed(),
        "index_state":   caen_index_state(),
        "stats":         svc.db_stats() if svc.is_indexed() else {},
        "known_missing": {f"{k[0]}_{k[1]}": v for k, v in Config.CAEN_KNOWN_MISSING.items()},
        "onrc_available": Config.ONRC_DB_PATH.exists() and get_onrc().is_indexed(),
    })


@app.post("/api/caen/start-index")
async def caen_start_index():
    svc = get_caen()
    if not svc.years_available():
        raise HTTPException(400, "No CAEN data files found in data/ folder")
    started = svc.start_indexing()
    return {"started": started, "message": "Indexing started" if started else "Already indexing"}


@app.get("/api/caen/index-progress")
async def caen_index_progress():
    return JSONResponse(caen_index_state())


@app.get("/api/caen/search")
async def caen_search(q: str = "", limit: int = 20):
    if not q or len(q) < 1:
        return JSONResponse([])
    return JSONResponse(get_caen().search_caen(q.strip(), min(limit, 50)))


class CAENAnalyzeReq(BaseModel):
    caen: int
    years: list[int]
    anchor_year: int | None = None
    peer_cui: int | None = None
    entity_types: list[str] | None = None  # None = all; ["commercial"] = commercial only


@app.post("/api/caen/analyze")
async def caen_analyze(req: CAENAnalyzeReq):
    svc = get_caen()
    if not svc.is_indexed():
        raise HTTPException(400, "CAEN data not indexed — run indexing first")

    # Fetch CAEN description from ANAF if not cached — non-critical, skip on any error
    if not svc.get_description(req.caen):
        sample_cui = svc.get_sample_cui(req.caen)
        if sample_cui:
            try:
                client = get_anaf_client()
                for year in reversed(req.years):
                    try:
                        fin_raw = await client.fetch_financials(sample_cui, year)
                        den = (fin_raw.get("den_caen") or "").strip()
                        if den:
                            svc.cache_description(req.caen, den)
                            break
                    except (asyncio.CancelledError, Exception):
                        continue
            except (asyncio.CancelledError, Exception) as ex:
                log.debug("CAEN desc fetch skipped: %s", ex)

    result = svc.analyze(
        caen=req.caen, years=req.years,
        anchor_year=req.anchor_year, peer_cui=req.peer_cui,
        entity_types=req.entity_types,
    )

    # Enrich top10 names from ANAF when ONRC not indexed
    top10 = result.get("top10_revenue") or []
    missing = [r["cui"] for r in top10 if r.get("name","").startswith("CUI ")]
    if missing:
        try:
            today  = datetime.date.today().strftime("%Y-%m-%d")
            client = get_anaf_client()
            raw    = await client.fetch_company_info(missing[:50], today)
            nm = {}
            for entry in raw.get("found") or []:
                dg = entry.get("date_generale") or {}
                c  = int(dg.get("cui",0) or 0)
                n  = (dg.get("denumire") or "").strip()
                if c and n:
                    nm[c] = n
                    svc.cache_company_name(c, n)
            for r in top10:
                if r.get("name","").startswith("CUI ") and r["cui"] in nm:
                    r["name"] = nm[r["cui"]]
        except Exception as ex:
            log.warning("CAEN top10 name fetch: %s", ex)

    return JSONResponse(content=result)


@app.get("/romania.geojson")
async def romania_geojson():
    from fastapi.responses import FileResponse as FR
    p = Path(__file__).parent.parent.parent / "romania.geojson"
    if not p.exists():
        raise HTTPException(404, "romania.geojson not found in project root")
    return FR(str(p), media_type="application/geo+json")


@app.post("/api/caen/export")
async def caen_export(req: CAENAnalyzeReq, eur: int = Query(default=0)):
    """Export full CAEN analysis as Excel workbook."""
    svc = get_caen()
    if not svc.is_indexed():
        raise HTTPException(400, "CAEN data not indexed")
    result = svc.analyze(
        caen=req.caen, years=req.years,
        anchor_year=req.anchor_year, peer_cui=req.peer_cui,
    )
    if result.get("empty"):
        raise HTTPException(404, result.get("reason", "No data"))

    eur_rates = Config.EUR_RATES if eur else None
    from backend.exporters.caen_exporter import export_caen_excel
    xlsx = export_caen_excel(result, req.years, eur_rates)
    cur  = "EUR" if eur else "RON"
    date = datetime.date.today().strftime("%Y%m%d")
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=caen_{req.caen}_{cur}_{date}.xlsx"},
    )


@app.post("/api/caen/export-full")
async def caen_export_full(req: CAENAnalyzeReq, eur: int = Query(default=0)):
    """Full export — every company in sector. May be large."""
    svc = get_caen()
    if not svc.is_indexed():
        raise HTTPException(400, "CAEN data not indexed")
    result = svc.analyze(
        caen=req.caen, years=req.years,
        anchor_year=req.anchor_year, peer_cui=req.peer_cui,
    )
    if result.get("empty"):
        raise HTTPException(404, result.get("reason", "No data"))

    all_rows = svc.get_all_rows(
        caen=req.caen, years=req.years, anchor_year=req.anchor_year
    )
    eur_rates = Config.EUR_RATES if eur else None
    from backend.exporters.caen_exporter import export_caen_full_excel
    xlsx = export_caen_full_excel(result, all_rows, req.years, eur_rates)
    cur  = "EUR" if eur else "RON"
    date = datetime.date.today().strftime("%Y%m%d")
    n    = len(all_rows)
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=caen_{req.caen}_full_{n}co_{cur}_{date}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════════════
# Name → CUI Matcher
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/match/template")
async def match_template():
    """Downloadable xlsx template — single column of company names."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Companii"
    ws.append(["nume_firma"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    for example in ["S.C. Exemplu Company S.R.L.", "Alta Firma SA", "A Treia Firma SRL"]:
        ws.append([example])
    ws.column_dimensions["A"].width = 42
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nume_firme_template.xlsx"},
    )


def _extract_names_from_upload(filename: str, content: bytes) -> tuple[list[str], list[str]]:
    """Parse xlsx or csv upload, return (names, warnings). Looks for a column
    named nume/nume_firma/company/denumire/name, else uses the first column."""
    warnings: list[str] = []
    names: list[str] = []
    lower = filename.lower()

    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], ["File is empty"]
        header = [str(h or "").strip().lower() for h in rows[0]]
        name_col_idx = 0
        for i, h in enumerate(header):
            if h in ("nume_firma", "nume", "company", "company_name", "denumire", "name"):
                name_col_idx = i
                break
        for rn, row in enumerate(rows[1:], 2):
            if name_col_idx >= len(row):
                continue
            val = row[name_col_idx]
            if val is None or not str(val).strip():
                continue
            names.append(str(val).strip())
    else:
        # CSV / TXT fallback
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return [], ["Cannot decode file (use UTF-8)"]
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], ["File is empty"]
        header = [h.strip().lower() for h in rows[0]]
        name_col_idx = 0
        for i, h in enumerate(header):
            if h in ("nume_firma", "nume", "company", "company_name", "denumire", "name"):
                name_col_idx = i
                break
        # If header row doesn't look like a header (no match and only 1 col), treat as data
        has_header = any(h in ("nume_firma", "nume", "company", "company_name", "denumire", "name") for h in header)
        data_rows = rows[1:] if has_header else rows
        for row in data_rows:
            if name_col_idx >= len(row):
                continue
            val = row[name_col_idx].strip()
            if val:
                names.append(val)

    if not names:
        warnings.append("No company names found in file")
    return names, warnings


@app.post("/api/match/start")
async def match_start(file: UploadFile = File(...)):
    onrc = get_onrc()
    if not onrc.is_indexed():
        raise HTTPException(400, "ONRC not indexed — index ONRC data first (required for name matching)")

    content = await file.read()
    names, warnings = _extract_names_from_upload(file.filename or "upload.csv", content)
    if not names:
        raise HTTPException(422, "; ".join(warnings) or "No valid company names found")
    if len(names) > 20000:
        raise HTTPException(422, f"Too many names ({len(names):,}) — maximum 20,000 per batch")

    mgr = get_match_manager()
    job = mgr.create(names)
    asyncio.create_task(run_match_job(job))
    return JSONResponse({
        "job_id": job.job_id, "total": job.total, "warnings": warnings[:10],
    })


@app.get("/api/match/status/{jid}")
async def match_status(jid: str):
    job = get_match_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    return JSONResponse(job.to_status_dict())


@app.get("/api/match/results/{jid}")
async def match_results(jid: str):
    job = get_match_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    if job.status not in ("done",):
        raise HTTPException(409, f"Job status is '{job.status}', not done yet")
    return JSONResponse({
        "job_id": job.job_id, "total": job.total, "summary": job.summary,
        "thresholds": {"auto": AUTO_THRESHOLD, "review": REVIEW_THRESHOLD},
        "results": [r.to_dict() for r in job.results],
    })


class MatchDecision(BaseModel):
    job_id: str
    action: str                    # "approve" | "reject" | "select" | "approve_all_above" | "reject_all_below" | "reset"
    row_index: int | None = None
    selected_cui: int | None = None
    threshold: float | None = None


@app.post("/api/match/decide")
async def match_decide(dec: MatchDecision):
    job = get_match_manager().get(dec.job_id)
    if not job: raise HTTPException(404, "Job not found")

    def _apply_select(r, cui: int):
        alt = next((a for a in r.alternatives if a.cui == cui), None)
        if alt:
            r.cui = alt.cui; r.matched_name = alt.denumire; r.score = alt.score
            r.judet = alt.judet; r.forma_juridica = alt.forma_juridica
        else:
            r.cui = cui
        r.status = "manual"

    if dec.action in ("approve", "reject", "select", "reset") and dec.row_index is not None:
        r = next((x for x in job.results if x.row_index == dec.row_index), None)
        if not r: raise HTTPException(404, "Row not found")
        if dec.action == "approve":
            r.status = "approved"
        elif dec.action == "reject":
            r.status = "rejected"; r.cui = None
        elif dec.action == "select" and dec.selected_cui:
            _apply_select(r, dec.selected_cui)
        elif dec.action == "reset":
            r.status = "pending" if r.score >= REVIEW_THRESHOLD or r.forced_review else "no_match"
        return {"ok": True, "row_index": r.row_index, "status": r.status}

    elif dec.action == "approve_all_above":
        th = dec.threshold if dec.threshold is not None else AUTO_THRESHOLD
        n = 0
        for r in job.results:
            if r.status == "pending" and r.score >= th:
                r.status = "approved"; n += 1
        return {"ok": True, "affected": n}

    elif dec.action == "reject_all_below":
        th = dec.threshold if dec.threshold is not None else REVIEW_THRESHOLD
        n = 0
        for r in job.results:
            if r.status == "pending" and r.score < th:
                r.status = "rejected"; r.cui = None; n += 1
        return {"ok": True, "affected": n}

    raise HTTPException(422, "Invalid action or missing parameters")


@app.get("/api/match/export/{jid}")
async def match_export(jid: str, fmt: str = Query(default="xlsx")):
    job = get_match_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    if job.status != "done":
        raise HTTPException(409, "Job not finished yet")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    headers = ["Nume Input", "Nume Găsit (ONRC)", "CUI", "Scor Similaritate",
               "Status", "Județ", "Formă Juridică"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G1"

    status_colors = {
        "auto":     "D1FAE5", "approved": "D1FAE5", "manual": "DBEAFE",
        "pending":  "FEF3C7", "rejected": "FEE2E2", "no_match": "F3F4F6",
    }
    status_labels = {
        "auto": "Auto-matched", "approved": "Approved (manual)", "manual": "Manually selected",
        "pending": "⚠ Pending review", "rejected": "✖ Rejected", "no_match": "No match found",
    }

    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in job.results:
        ws.append([
            r.input_name, r.matched_name or "", r.cui or "",
            round(r.score, 1) if r.score else "",
            status_labels.get(r.status, r.status), r.judet, r.forma_juridica,
        ])
        row_i = ws.max_row
        fill = PatternFill("solid", fgColor=status_colors.get(r.status, "FFFFFF"))
        for c in ws[row_i]:
            c.fill = fill; c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center")

    widths = [38, 38, 14, 12, 18, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    s = job.summary
    note_row = len(job.results) + 3
    note = ws.cell(row=note_row, column=1,
                   value=(f"{s['auto']+s['approved']+s['manual']:,} matched · "
                          f"{s['pending']:,} pending · {s['rejected']+s['no_match']:,} no match — "
                          f"Exported {datetime.date.today()}"))
    note.font = Font(italic=True, color="666666", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date = datetime.date.today().strftime("%Y%m%d")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=name_match_{jid}_{date}.xlsx"},
    )


@app.get("/api/match/to-bulk/{jid}")
async def match_to_bulk(jid: str):
    """Returns the confirmed CUIs (auto + approved + manual) ready to feed into Bulk Analysis."""
    job = get_match_manager().get(jid)
    if not job: raise HTTPException(404, "Job not found")
    cuis = [r.cui for r in job.results if r.status in ("auto", "approved", "manual") and r.cui]
    return JSONResponse({"cuis": cuis, "count": len(cuis)})
